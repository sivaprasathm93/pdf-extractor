import os
import io
import re
import tempfile
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import fitz  # PyMuPDF
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from analytics import track_event, get_stats, get_recent_events

app = Flask(__name__)
CORS(app)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max upload

UPLOAD_FOLDER = tempfile.mkdtemp()


class PasswordRequiredError(Exception):
    """Raised when a PDF requires a password to open."""
    pass


# Kotak date format: "01 Jan, 2026"
DATE_PATTERN = re.compile(
    r'(\d{1,2}\s+(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec),?\s*\d{4})',
    re.IGNORECASE
)
# DD-MM-YYYY and DD/MM/YYYY
DATE_PATTERN_ALT = re.compile(r'(\d{2}[-/]\d{2}[-/]\d{2,4})')

# Transaction type prefixes used to identify the start of a narration block
TXN_PREFIXES = re.compile(
    r'^(IMPS|UPI|NEFT|RTGS|POS|FT|ACH|ECS|ATM|CHQ|CLG|DD|IFT|MOB|NET|SI|NACH|BIL|EMI)',
    re.IGNORECASE
)


def parse_bank_statement(pdf_path, password=None):
    """Extract transaction table data from bank statement PDF using PyMuPDF."""
    all_transactions = []
    doc = fitz.open(pdf_path)

    # Handle password-protected PDFs
    if doc.is_encrypted:
        if not password:
            doc.close()
            raise PasswordRequiredError("PDF is password-protected")
        if not doc.authenticate(password):
            doc.close()
            raise ValueError("Incorrect PDF password")

    for page_num in range(len(doc)):
        page = doc[page_num]
        page_txns = []

        # Try table extraction first (PyMuPDF built-in table finder)
        try:
            tabs = page.find_tables()
            if tabs and tabs.tables:
                for table in tabs.tables:
                    rows = table.extract()
                    header_idx = find_header_row(rows)
                    if header_idx >= 0:
                        col_map = map_columns(rows[header_idx])
                        for row in rows[header_idx + 1:]:
                            # Handle multi-line cells (HDFC format: all txns in one row)
                            split_txns = split_multiline_row(row, col_map)
                            if split_txns:
                                page_txns.extend(split_txns)
                            else:
                                txn = parse_table_row(row, col_map)
                                if txn:
                                    page_txns.append(txn)
        except Exception:
            pass

        # If table extraction found data, use it; otherwise try text-based
        if page_txns:
            all_transactions.extend(page_txns)
        else:
            # Fallback: text-based extraction
            text = page.get_text("text")
            if text:
                text_txns = extract_from_text(text)
                all_transactions.extend(text_txns)

    doc.close()

    # Deduplicate
    seen = set()
    unique = []
    for txn in all_transactions:
        key = (txn["date"], txn["details"][:30], txn["balance"])
        if key not in seen:
            seen.add(key)
            unique.append(txn)

    return unique


def split_multiline_row(row, col_map):
    """
    Handle table rows where ALL transactions are merged into one row
    with newline characters inside cells (common in HDFC statements).
    
    Dates/amounts have 1 line per transaction, but narration has multiple
    lines per transaction (typically 3 for HDFC).
    """
    date_idx = col_map.get("date")
    if date_idx is None or date_idx >= len(row):
        return None

    date_cell = str(row[date_idx]) if row[date_idx] else ""
    if "\n" not in date_cell:
        return None

    # Split date column and find valid dates
    date_lines = [d.strip() for d in date_cell.split("\n") if d.strip()]
    valid_dates = [d for d in date_lines if is_date(d)]
    if len(valid_dates) <= 1:
        return None

    num_txns = len(valid_dates)

    # Split amount columns — should have same count as dates
    def split_amount_col(key):
        idx = col_map.get(key)
        if idx is None or idx >= len(row):
            return [""] * num_txns
        cell_val = str(row[idx]) if row[idx] else ""
        lines = [l.strip() for l in cell_val.split("\n") if l.strip()]
        # Pad to match transaction count
        while len(lines) < num_txns:
            lines.append("")
        return lines[:num_txns]

    debit_lines = split_amount_col("debit")
    credit_lines = split_amount_col("credit")
    balance_lines = split_amount_col("balance")
    cheque_lines = split_amount_col("cheque")

    # Split narration — may have MORE lines than dates
    details_idx = col_map.get("details")
    if details_idx is not None and details_idx < len(row):
        details_cell = str(row[details_idx]) if row[details_idx] else ""
        details_all_lines = [l.strip() for l in details_cell.split("\n") if l.strip()]
    else:
        details_all_lines = [""] * num_txns

    # Group narration lines into N transactions
    narration_groups = group_narration_lines(details_all_lines, num_txns)

    transactions = []
    for i in range(num_txns):
        date_val = valid_dates[i]
        details_val = narration_groups[i] if i < len(narration_groups) else ""

        debit_val = clean_amount(debit_lines[i])
        credit_val = clean_amount(credit_lines[i])
        balance_val = clean_amount(balance_lines[i])
        cheque_val = cheque_lines[i] if i < len(cheque_lines) else ""

        if "OPENING BALANCE" in details_val.upper() or "CLOSING BALANCE" in details_val.upper():
            continue

        transactions.append({
            "date": date_val.strip(),
            "details": details_val.strip(),
            "cheque": cheque_val.strip(),
            "debit": debit_val,
            "credit": credit_val,
            "balance": balance_val,
        })

    return transactions if transactions else None


def group_narration_lines(lines, num_groups):
    """
    Group narration lines into the correct number of transactions.
    
    HDFC narration end markers (in priority order):
      1. Line containing 'Ref <digits>' (e.g., '01/01/2026 Ref 600114982784')
      2. Standalone digits line after line ending with 'Ref'
         (e.g., 'hekam Value Dt 03/01/2026 Ref' + '600335454148')
      3. Line containing 'Value Dt DD/MM/YYYY' without Ref
         (e.g., 'Value Dt 06/01/2026')
    """
    if not lines:
        return [""] * num_groups

    if len(lines) == num_groups:
        return lines[:]

    if len(lines) < num_groups:
        result = lines[:]
        while len(result) < num_groups:
            result.append("")
        return result

    # --- More lines than groups: need to group them ---

    # Mark each line as "end of narration" or not
    ref_same_line = re.compile(r'Ref\s+\d+', re.IGNORECASE)
    ref_end_of_line = re.compile(r'Ref\s*$', re.IGNORECASE)
    standalone_digits = re.compile(r'^\d{6,}$')
    value_dt_pattern = re.compile(r'Value\s+Dt\s+\d{2}/\d{2}/\d{4}', re.IGNORECASE)

    end_indices = []
    skip_next = False
    for i, line in enumerate(lines):
        if skip_next:
            skip_next = False
            continue

        stripped = line.strip()

        # Case 1: 'Ref 600114982784' on same line
        if ref_same_line.search(stripped):
            end_indices.append(i)
            continue

        # Case 2: line ends with 'Ref', next line is standalone digits
        if ref_end_of_line.search(stripped) and i + 1 < len(lines):
            next_stripped = lines[i + 1].strip()
            if standalone_digits.match(next_stripped):
                end_indices.append(i + 1)
                skip_next = True
                continue

        # Case 3: 'Value Dt DD/MM/YYYY' without Ref (end of narration)
        # Only match if this line contains Value Dt AND does NOT contain 'Ref'
        if value_dt_pattern.search(stripped) and 'ref' not in stripped.lower():
            end_indices.append(i)
            continue

    if len(end_indices) >= num_groups:
        end_indices = end_indices[:num_groups]
        groups = []
        start = 0
        for end_idx in end_indices:
            group_text = " ".join(lines[start:end_idx + 1])
            groups.append(group_text)
            start = end_idx + 1
        while len(groups) < num_groups:
            groups.append("")
        return groups

    # Fallback: even distribution
    lines_per_group = len(lines) / num_groups
    groups = []
    for i in range(num_groups):
        start = int(round(i * lines_per_group))
        end = int(round((i + 1) * lines_per_group))
        group_text = " ".join(lines[start:end])
        groups.append(group_text)
    return groups


def find_header_row(rows):
    """Find the header row index in a table."""
    for i, row in enumerate(rows):
        if not row:
            continue
        row_text = " ".join(str(cell) if cell else "" for cell in row).upper()
        if "DATE" in row_text and ("DEBIT" in row_text or "CREDIT" in row_text or
                                    "BALANCE" in row_text or "WITHDRAWAL" in row_text or
                                    "DEPOSIT" in row_text):
            return i
    return -1


def map_columns(headers):
    """Map header names to column indices."""
    col_map = {}
    for i, h in enumerate(headers):
        if not h:
            continue
        h_upper = str(h).upper().strip()
        if "DATE" in h_upper and "date" not in col_map:
            col_map["date"] = i
        elif any(kw in h_upper for kw in ["TRANSACTION", "NARRATION", "DESCRIPTION", "DETAIL", "PARTICULAR"]):
            col_map["details"] = i
        elif "CHEQUE" in h_upper or "CHQ" in h_upper or "REF" in h_upper:
            col_map["cheque"] = i
        elif "DEBIT" in h_upper or "WITHDRAWAL" in h_upper or h_upper == "DR":
            col_map["debit"] = i
        elif "CREDIT" in h_upper or "DEPOSIT" in h_upper or h_upper == "CR":
            col_map["credit"] = i
        elif "BALANCE" in h_upper or "BAL" in h_upper:
            col_map["balance"] = i
    return col_map


def parse_table_row(row, col_map):
    """Parse a single table row into a transaction dict."""
    date_val = get_col_value(row, col_map.get("date"))
    # Handle cells with newlines — take the first line that looks like a date
    if "\n" in date_val:
        for part in date_val.split("\n"):
            if is_date(part.strip()):
                date_val = part.strip()
                break

    if not date_val or not is_date(date_val):
        return None

    details = get_col_value(row, col_map.get("details"))
    if "OPENING BALANCE" in details.upper() or "CLOSING BALANCE" in details.upper():
        return None

    debit = clean_amount(get_col_value(row, col_map.get("debit")))
    credit = clean_amount(get_col_value(row, col_map.get("credit")))

    return {
        "date": date_val.strip(),
        "details": details.replace("\n", " ").strip(),
        "cheque": get_col_value(row, col_map.get("cheque")).replace("\n", " ").strip(),
        "debit": debit,
        "credit": credit,
        "balance": clean_amount(get_col_value(row, col_map.get("balance"))),
    }


def extract_from_text(text):
    """Fallback text-based extraction for pages without clear table structure."""
    transactions = []
    lines = text.split('\n')

    full_text = text.upper()
    if "DATE" not in full_text or ("DEBIT" not in full_text and "CREDIT" not in full_text and
                                     "WITHDRAWAL" not in full_text and "DEPOSIT" not in full_text):
        return transactions

    i = 0
    while i < len(lines):
        line = lines[i].strip()
        match = DATE_PATTERN.match(line)
        if not match:
            match = DATE_PATTERN_ALT.match(line)

        if match:
            date_val = match.group(1)
            rest = line[match.end():].strip()

            details_parts = []
            cheque = ""
            nums = []

            if rest:
                parts = rest.split()
                for part in parts:
                    cleaned_part = part.replace(",", "").replace("+", "").replace("-", "").strip()
                    try:
                        float(cleaned_part)
                        nums.append(part)
                    except ValueError:
                        details_parts.append(part)

            j = i + 1
            while j < len(lines):
                next_line = lines[j].strip()
                if not next_line:
                    j += 1
                    continue

                if DATE_PATTERN.match(next_line) or DATE_PATTERN_ALT.match(next_line):
                    break

                if next_line.startswith("UPI-") or next_line.startswith("IMPS-") or re.match(r'^\d{12,}$', next_line):
                    cheque = next_line
                    j += 1
                    continue

                cleaned = next_line.replace(",", "").replace("+", "").replace("-", "").strip()
                try:
                    float(cleaned)
                    nums.append(next_line.strip())
                    j += 1
                    continue
                except ValueError:
                    pass

                if re.match(r'Page \d+ of \d+', next_line) or next_line in ["SUMMARY", "AP-Auto", "AP-Aut"]:
                    break

                details_parts.append(next_line)
                j += 1

            i = j

            details = " ".join(details_parts)

            if "OPENING BALANCE" in details.upper() or "CLOSING BALANCE" in details.upper():
                continue
            if "HOME BRANCH" in details.upper() or "SPENDZ" in details.upper():
                continue
            if not details and not nums:
                continue
            if details.startswith("-") and DATE_PATTERN.search(details):
                continue

            debit = ""
            credit = ""
            balance = ""

            if len(nums) >= 1:
                balance = clean_amount(nums[-1])
            if len(nums) >= 2:
                amount = nums[-2].strip()
                if amount.startswith("-"):
                    debit = clean_amount(amount)
                elif amount.startswith("+"):
                    credit = clean_amount(amount)
                else:
                    debit = clean_amount(amount)

            if details or debit or credit:
                transactions.append({
                    "date": date_val,
                    "details": details,
                    "cheque": cheque,
                    "debit": debit,
                    "credit": credit,
                    "balance": balance,
                })
        else:
            i += 1

    return transactions


def clean_amount(val):
    """Clean an amount string."""
    if not val:
        return ""
    val = val.strip()
    if val.startswith("+") or val.startswith("-"):
        val = val[1:]
    return val


def parse_float(val):
    """Convert a string amount to float for Excel. Returns None if empty."""
    if not val or not val.strip():
        return None
    try:
        return float(val.strip().replace(",", ""))
    except (ValueError, TypeError):
        return val


def get_col_value(row, idx):
    """Safely get a column value from a row."""
    if idx is None or idx >= len(row):
        return ""
    val = row[idx]
    return str(val) if val else ""


def is_date(text):
    """Check if text looks like a date (supports multiple formats)."""
    text = str(text).strip()
    if DATE_PATTERN.match(text):
        return True
    if DATE_PATTERN_ALT.match(text):
        return True
    return False


# ─── API Routes ──────────────────────────────────────────────────────────────


@app.route("/health", methods=["GET"])
def health():
    """Keep-alive health check endpoint."""
    return jsonify({"status": "ok"}), 200

@app.route("/api/debug", methods=["POST"])
def debug_pdf():
    """Debug endpoint: show raw table and text extraction for diagnosis."""
    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400

    file = request.files["file"]
    password = request.form.get("password", "").strip() or None

    filepath = os.path.join(UPLOAD_FOLDER, "debug.pdf")
    file.save(filepath)

    doc = fitz.open(filepath)
    if doc.is_encrypted:
        if password:
            doc.authenticate(password)
        else:
            doc.close()
            return jsonify({"error": "PDF is encrypted, provide password"}), 400

    result = {"pages": []}

    for page_num in range(len(doc)):
        page = doc[page_num]
        page_data = {"page": page_num + 1, "tables": [], "text_lines": []}

        try:
            tabs = page.find_tables()
            if tabs and tabs.tables:
                for t_idx, table in enumerate(tabs.tables):
                    rows = table.extract()
                    page_data["tables"].append({
                        "table_index": t_idx,
                        "row_count": len(rows),
                        "rows": [[str(cell) if cell else "" for cell in row] for row in rows[:30]],
                    })
        except Exception as e:
            page_data["table_error"] = str(e)

        text = page.get_text("text")
        if text:
            page_data["text_lines"] = text.split("\n")[:80]

        result["pages"].append(page_data)

    doc.close()
    return jsonify(result)


@app.route("/api/upload", methods=["POST"])
def upload_pdf():
    """Upload a PDF and extract transaction data."""
    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400

    file = request.files["file"]
    if not file.filename.lower().endswith(".pdf"):
        return jsonify({"error": "Only PDF files are allowed"}), 400

    password = request.form.get("password", "").strip() or None

    filepath = os.path.join(UPLOAD_FOLDER, "statement.pdf")
    file.save(filepath)

    try:
        transactions = parse_bank_statement(filepath, password=password)

        response_data = {
            "success": True,
            "data": transactions,
            "count": len(transactions),
            "filename": file.filename,
        }

        # If PDF was encrypted, save a decrypted copy for preview
        if password:
            try:
                doc = fitz.open(filepath)
                if doc.is_encrypted:
                    doc.authenticate(password)
                decrypted_path = os.path.join(UPLOAD_FOLDER, "preview.pdf")
                doc.save(decrypted_path)
                doc.close()
                response_data["preview_url"] = f"{request.host_url}api/preview"
            except Exception:
                pass

        return jsonify(response_data)
    except PasswordRequiredError:
        return jsonify({
            "password_required": True,
            "error": "This PDF is password-protected. Please provide the password.",
        }), 200
    except ValueError as e:
        return jsonify({
            "password_required": True,
            "error": str(e),
        }), 200
    except Exception as e:
        return jsonify({"error": f"Failed to parse PDF: {str(e)}"}), 500


@app.route("/api/preview-upload", methods=["POST"])
def preview_upload():
    """Upload PDF for preview, decrypting if needed."""
    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400

    file = request.files["file"]
    password = request.form.get("password", "")
    filepath = os.path.join(UPLOAD_FOLDER, "statement.pdf")
    file.save(filepath)

    try:
        doc = fitz.open(filepath)
        if doc.is_encrypted:
            if not password:
                doc.close()
                return jsonify({
                    "password_required": True,
                    "error": "This PDF is password-protected. Please enter the password.",
                }), 200
            if not doc.authenticate(password):
                doc.close()
                return jsonify({
                    "password_required": True,
                    "error": "Incorrect password. Please try again.",
                }), 200
            # Save decrypted copy
            decrypted_path = os.path.join(UPLOAD_FOLDER, "preview.pdf")
            doc.save(decrypted_path)
            doc.close()
            return jsonify({
                "success": True,
                "preview_url": f"{request.host_url}api/preview",
            })
        else:
            doc.close()
            return jsonify({
                "success": True,
                "preview_url": f"{request.host_url}api/preview",
            })
    except Exception as e:
        return jsonify({"error": f"Failed to open PDF: {str(e)}"}), 500


@app.route("/api/preview", methods=["GET"])
def preview_pdf():
    """Serve the decrypted PDF for preview."""
    decrypted_path = os.path.join(UPLOAD_FOLDER, "preview.pdf")
    original_path = os.path.join(UPLOAD_FOLDER, "statement.pdf")

    filepath = decrypted_path if os.path.exists(decrypted_path) else original_path
    if not os.path.exists(filepath):
        return jsonify({"error": "No PDF available"}), 404

    return send_file(filepath, mimetype="application/pdf")


@app.route("/api/export", methods=["POST"])
def export_excel():
    """Export transaction data to Excel."""
    data = request.json
    if not data or "transactions" not in data:
        return jsonify({"error": "No data provided"}), 400

    transactions = data["transactions"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Bank Statement"

    # Header styling
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Write headers
    headers = ["Date", "Transaction Details", "Cheque/Ref No.", "Debit", "Credit", "Balance"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Write data
    data_font = Font(name="Calibri", size=10)
    number_format = '#,##0.00'
    for row_idx, txn in enumerate(transactions, 2):
        values = [
            txn.get("date", ""),
            txn.get("details", ""),
            txn.get("cheque", ""),
            parse_float(txn.get("debit", "")),
            parse_float(txn.get("credit", "")),
            parse_float(txn.get("balance", "")),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = data_font
            cell.border = thin_border
            if col >= 4:
                cell.alignment = Alignment(horizontal="right")
                if val is not None and isinstance(val, (int, float)):
                    cell.number_format = number_format

    # Column widths
    col_widths = [16, 50, 20, 15, 15, 15]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="bank_statement.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ─── Analytics Routes ─────────────────────────────────────────────────────────


@app.route("/api/track", methods=["POST"])
def track():
    """Record an analytics event."""
    data = request.json
    if not data or "event" not in data:
        return jsonify({"error": "Missing event field"}), 400

    event_type = data["event"]
    allowed = ["page_visit", "extract_click", "export_click", "preview_click"]
    if event_type not in allowed:
        return jsonify({"error": f"Unknown event type: {event_type}"}), 400

    ip = request.headers.get("X-Forwarded-For", request.remote_addr)
    ua = request.headers.get("User-Agent", "")
    track_event(event_type, ip_address=ip, user_agent=ua)

    return jsonify({"success": True}), 200


@app.route("/api/analytics", methods=["GET"])
def analytics():
    """Get aggregate analytics stats."""
    return jsonify(get_stats())


@app.route("/api/analytics/events", methods=["GET"])
def analytics_events():
    """Get recent analytics events."""
    limit = request.args.get("limit", 50, type=int)
    return jsonify(get_recent_events(limit))


if __name__ == "__main__":
    app.run(debug=True, port=5000)
